"""
Sensor Map Excel generation service.

Responsibilities
----------------
1. Read Sensor Map section rules from config/sensormap_sections.json.
2. Parse the project's Peripheral Sensor Scope.
3. Copy the configured Excel template to a target directory.
4. Locate Sensor sections dynamically by their configured title text.
5. Keep sections present in the scope and delete sections not present.
6. Save the generated workbook and return a structured execution result.

This module does not depend on FastAPI. The API layer can call
`generate_sensor_map(...)` later.
"""

from __future__ import annotations

import json
import logging
import re
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet


LOGGER = logging.getLogger(__name__)

DEFAULT_CONFIG_PATH = (
    Path(__file__).resolve().parents[1]
    / "config"
    / "sensormap_sections.json"
)


class SensorMapError(RuntimeError):
    """Base exception for Sensor Map generation failures."""


class SensorMapConfigError(SensorMapError):
    """Raised when the Sensor Map JSON configuration is invalid."""


class SensorMapTemplateError(SensorMapError):
    """Raised when the Excel template cannot be found or opened."""


class SensorMapSectionError(SensorMapError):
    """Raised when configured Sensor sections cannot be resolved safely."""


@dataclass(frozen=True)
class SectionLocation:
    """A Sensor section located in a worksheet."""

    sensor_code: str
    start_row: int
    end_row: int
    title: str


def _normalize_text(value: Any) -> str:
    """
    Normalize text for reliable comparisons.

    Normalization includes:
    - conversion to string;
    - replacement of full-width punctuation commonly found in templates;
    - trimming;
    - collapsing repeated whitespace;
    - case-insensitive comparison through casefold().
    """
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\u3000", " ")
    text = text.replace("：", ":")
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", " ", text).strip()
    return text.casefold()


def load_sensormap_config(
    config_path: str | Path = DEFAULT_CONFIG_PATH,
) -> dict[str, Any]:
    """Load and validate the Sensor Map JSON configuration."""
    path = Path(config_path)

    if not path.is_file():
        raise SensorMapConfigError(
            f"Sensor Map config file does not exist: {path}"
        )

    try:
        with path.open("r", encoding="utf-8") as file:
            config = json.load(file)
    except json.JSONDecodeError as exc:
        raise SensorMapConfigError(
            f"Sensor Map config is not valid JSON: {path}. "
            f"Line {exc.lineno}, column {exc.colno}: {exc.msg}"
        ) from exc
    except OSError as exc:
        raise SensorMapConfigError(
            f"Unable to read Sensor Map config: {path}"
        ) from exc

    if not isinstance(config, dict):
        raise SensorMapConfigError(
            "Sensor Map config root must be a JSON object."
        )

    template_path = config.get("template_path")
    if not isinstance(template_path, str) or not template_path.strip():
        raise SensorMapConfigError(
            '"template_path" must be a non-empty string.'
        )

    sensor_sections = config.get("sensor_sections")
    if not isinstance(sensor_sections, dict) or not sensor_sections:
        raise SensorMapConfigError(
            '"sensor_sections" must be a non-empty JSON object.'
        )

    for code, rule in sensor_sections.items():
        if not isinstance(rule, dict):
            raise SensorMapConfigError(
                f'Sensor rule "{code}" must be a JSON object.'
            )

        aliases = rule.get("aliases")
        titles = rule.get("titles")

        if not isinstance(aliases, list) or not any(
            isinstance(item, str) and item.strip() for item in aliases
        ):
            raise SensorMapConfigError(
                f'Sensor rule "{code}" must contain a non-empty '
                '"aliases" string list.'
            )

        if not isinstance(titles, list) or not any(
            isinstance(item, str) and item.strip() for item in titles
        ):
            raise SensorMapConfigError(
                f'Sensor rule "{code}" must contain a non-empty '
                '"titles" string list.'
            )

    return config


def parse_peripheral_sensor_scope(
    scope: str | None,
    sensor_sections: Mapping[str, Mapping[str, Any]],
) -> set[str]:
    """
    Convert a Peripheral Sensor Scope string to configured Sensor codes.

    Examples:
        "2*UFS6s+2*PAS6s+2*PPS3" -> {"UFS", "PAS", "PPS"}
        "PCS + RCS"               -> {"PCS", "RCS"}

    Matching is case-insensitive. Aliases may be followed or preceded by
    digits, but are not matched as part of a longer alphabetic word.
    """
    normalized_scope = _normalize_text(scope)
    if not normalized_scope:
        return set()

    detected: set[str] = set()

    for raw_code, rule in sensor_sections.items():
        code = str(raw_code).upper().strip()
        aliases = rule.get("aliases", [])

        for alias in aliases:
            normalized_alias = _normalize_text(alias)
            if not normalized_alias:
                continue

            pattern = (
                rf"(?<![a-z])"
                rf"{re.escape(normalized_alias)}"
                rf"(?![a-z])"
            )
            if re.search(pattern, normalized_scope):
                detected.add(code)
                break

    return detected


def _iter_non_empty_cells(ws: Worksheet) -> Iterable[tuple[int, int, str]]:
    """Yield row, column and normalized text for all non-empty cells."""
    for row in ws.iter_rows():
        for cell in row:
            normalized = _normalize_text(cell.value)
            if normalized:
                yield cell.row, cell.column, normalized


def _find_title_rows(
    ws: Worksheet,
    sensor_sections: Mapping[str, Mapping[str, Any]],
) -> dict[str, tuple[int, str]]:
    """
    Find each configured Sensor title in the worksheet.

    A configured Sensor must occur exactly once. Duplicate matches are
    rejected because deleting an ambiguous range could corrupt the template.
    """
    title_lookup: dict[str, list[tuple[str, str]]] = {}

    for raw_code, rule in sensor_sections.items():
        code = str(raw_code).upper().strip()
        for title in rule.get("titles", []):
            normalized = _normalize_text(title)
            if normalized:
                title_lookup.setdefault(normalized, []).append((code, title))

    matches: dict[str, list[tuple[int, int, str]]] = {
        str(code).upper().strip(): [] for code in sensor_sections
    }

    for row, column, cell_text in _iter_non_empty_cells(ws):
        candidates = title_lookup.get(cell_text, [])
        for code, configured_title in candidates:
            matches[code].append((row, column, configured_title))

    result: dict[str, tuple[int, str]] = {}
    errors: list[str] = []

    for code, locations in matches.items():
        if not locations:
            expected = sensor_sections[code].get("titles", [])
            errors.append(
                f'{code}: title not found; expected one of {expected}'
            )
            continue

        unique_rows = sorted({row for row, _, _ in locations})
        if len(unique_rows) > 1:
            errors.append(
                f"{code}: title found on multiple rows {unique_rows}"
            )
            continue

        result[code] = (locations[0][0], locations[0][2])

    if errors:
        raise SensorMapSectionError(
            "Unable to locate Sensor sections safely: " + "; ".join(errors)
        )

    return result


def _find_end_marker_row(
    ws: Worksheet,
    configured_markers: Sequence[str],
    after_row: int,
) -> int | None:
    """Find the first configured section-end marker below `after_row`."""
    markers = {
        _normalize_text(marker)
        for marker in configured_markers
        if isinstance(marker, str) and marker.strip()
    }
    if not markers:
        return None

    rows: list[int] = []
    for row, _, cell_text in _iter_non_empty_cells(ws):
        if row > after_row and cell_text in markers:
            rows.append(row)

    return min(rows) if rows else None


def find_sensor_sections(
    ws: Worksheet,
    config: Mapping[str, Any],
) -> dict[str, SectionLocation]:
    """
    Dynamically calculate Sensor section ranges.

    A section starts at its configured title row and ends immediately before
    the next Sensor title. The last Sensor section ends immediately before the
    first configured `section_end_markers` title.
    """
    sensor_sections = config["sensor_sections"]
    title_rows = _find_title_rows(ws, sensor_sections)

    ordered = sorted(
        (
            row,
            code,
            title,
        )
        for code, (row, title) in title_rows.items()
    )

    if len({row for row, _, _ in ordered}) != len(ordered):
        raise SensorMapSectionError(
            "Two configured Sensor sections resolve to the same row."
        )

    last_sensor_start = ordered[-1][0]
    end_marker_row = _find_end_marker_row(
        ws,
        config.get("section_end_markers", []),
        after_row=last_sensor_start,
    )

    if end_marker_row is None:
        raise SensorMapSectionError(
            "No configured section end marker was found after the final "
            "Sensor section. Add the next non-Sensor heading to "
            '"section_end_markers" in sensormap_sections.json.'
        )

    result: dict[str, SectionLocation] = {}

    for index, (start_row, code, title) in enumerate(ordered):
        if index + 1 < len(ordered):
            end_row = ordered[index + 1][0] - 1
        else:
            end_row = end_marker_row - 1

        if end_row < start_row:
            raise SensorMapSectionError(
                f"Invalid calculated range for {code}: "
                f"{start_row}-{end_row}"
            )

        result[code] = SectionLocation(
            sensor_code=code,
            start_row=start_row,
            end_row=end_row,
            title=title,
        )

    return result


def _row_is_in_deleted_range(
    row: int,
    delete_ranges: Sequence[tuple[int, int]],
) -> bool:
    return any(start <= row <= end for start, end in delete_ranges)


def _deleted_rows_before(
    row: int,
    delete_ranges: Sequence[tuple[int, int]],
) -> int:
    count = 0
    for start, end in delete_ranges:
        if end < row:
            count += end - start + 1
        elif start < row <= end:
            count += row - start
    return count


def _shift_row_after_deletions(
    row: int,
    delete_ranges: Sequence[tuple[int, int]],
) -> int:
    return row - _deleted_rows_before(row, delete_ranges)


def _capture_and_clear_merged_ranges(
    ws: Worksheet,
) -> list[tuple[int, int, int, int]]:
    """
    Capture all merged-cell coordinates and temporarily unmerge them.

    openpyxl does not reliably update merged-cell coordinates after deleting
    worksheet rows. Rebuilding the ranges after deletion prevents stale merge
    definitions.
    """
    ranges: list[tuple[int, int, int, int]] = []

    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(
            str(merged_range)
        )
        ranges.append((min_col, min_row, max_col, max_row))
        ws.unmerge_cells(str(merged_range))

    return ranges


def _restore_merged_ranges(
    ws: Worksheet,
    original_ranges: Sequence[tuple[int, int, int, int]],
    delete_ranges: Sequence[tuple[int, int]],
) -> None:
    """
    Restore merged ranges after row deletion.

    Merges fully inside a removed section are discarded. Merges crossing a
    deleted section boundary are rejected because automatically shrinking
    them could silently produce an invalid report.
    """
    for min_col, min_row, max_col, max_row in original_ranges:
        deleted_rows = [
            row
            for row in range(min_row, max_row + 1)
            if _row_is_in_deleted_range(row, delete_ranges)
        ]

        if deleted_rows:
            if len(deleted_rows) == (max_row - min_row + 1):
                continue
            raise SensorMapSectionError(
                "A merged cell crosses a Sensor section boundary: "
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max_row}. "
                "Adjust the template so merged ranges stay inside one section."
            )

        new_min_row = _shift_row_after_deletions(
            min_row,
            delete_ranges,
        )
        new_max_row = _shift_row_after_deletions(
            max_row,
            delete_ranges,
        )

        ws.merge_cells(
            start_row=new_min_row,
            start_column=min_col,
            end_row=new_max_row,
            end_column=max_col,
        )


def delete_unused_sensor_sections(
    ws: Worksheet,
    sections: Mapping[str, SectionLocation],
    sensors_to_keep: set[str],
) -> list[str]:
    """
    Delete all configured Sensor sections not present in `sensors_to_keep`.

    Deletion is performed from bottom to top so earlier row numbers remain
    valid. Merged ranges are rebuilt after deletion.
    """
    unsupported = sorted(sensors_to_keep - set(sections))
    if unsupported:
        raise SensorMapSectionError(
            "Scope contains configured Sensor codes that were not found in "
            f"the worksheet: {unsupported}"
        )

    sections_to_remove = [
        section
        for code, section in sections.items()
        if code not in sensors_to_keep
    ]

    if not sections_to_remove:
        return []

    delete_ranges = sorted(
        (
            (section.start_row, section.end_row)
            for section in sections_to_remove
        ),
        key=lambda item: item[0],
    )

    original_merges = _capture_and_clear_merged_ranges(ws)

    for start_row, end_row in reversed(delete_ranges):
        ws.delete_rows(
            start_row,
            end_row - start_row + 1,
        )

    _restore_merged_ranges(
        ws,
        original_merges,
        delete_ranges,
    )

    return sorted(section.sensor_code for section in sections_to_remove)


def _select_worksheet(workbook: Any, worksheet_name: str | None) -> Worksheet:
    """Select the configured worksheet, or the active sheet when blank."""
    requested = str(worksheet_name or "").strip()
    if not requested:
        return workbook.active

    if requested not in workbook.sheetnames:
        raise SensorMapTemplateError(
            f'Worksheet "{requested}" does not exist. '
            f"Available worksheets: {workbook.sheetnames}"
        )

    return workbook[requested]


def _resolve_output_filename(
    template_path: Path,
    output_filename: str | None,
    project_name: str | None,
) -> str:
    """Build a safe output filename while preserving the template suffix."""
    suffix = template_path.suffix.lower()
    requested = str(output_filename or "").strip()

    if requested:
        candidate = Path(requested).name
        if not candidate.lower().endswith(suffix):
            candidate += suffix
        return candidate

    safe_project = re.sub(
        r'[<>:"/\\|?*\x00-\x1f]+',
        "_",
        str(project_name or "Sensor_Map"),
    ).strip(" ._")

    if not safe_project:
        safe_project = "Sensor_Map"

    date_text = datetime.now().strftime("%Y%m%d")
    return f"{safe_project}_Sensor_Map_{date_text}{suffix}"


def generate_sensor_map(
    peripheral_sensor_scope: str | None,
    output_directory: str | Path,
    *,
    project_name: str | None = None,
    output_filename: str | None = None,
    config_path: str | Path = DEFAULT_CONFIG_PATH,
    overwrite: bool = True,
) -> dict[str, Any]:
    """
    Generate a project-specific Sensor Map workbook.

    Parameters
    ----------
    peripheral_sensor_scope:
        Current project's Peripheral Sensor Scope, for example
        ``"2*UFS6s+2*PAS6s+2*PPS3"``.
    output_directory:
        Final project Public Link Sensor Map directory.
    project_name:
        Optional project name used in the generated filename.
    output_filename:
        Optional explicit filename.
    config_path:
        Path to sensormap_sections.json.
    overwrite:
        Replace an existing output workbook when True.

    Returns
    -------
    dict
        Structured details including output path, kept sections and removed
        sections.
    """
    config = load_sensormap_config(config_path)

    template_path = Path(config["template_path"]).expanduser()
    if not template_path.is_file():
        raise SensorMapTemplateError(
            f"Sensor Map template does not exist: {template_path}"
        )

    supported_suffixes = {".xlsx", ".xlsm"}
    if template_path.suffix.lower() not in supported_suffixes:
        raise SensorMapTemplateError(
            "Sensor Map template must be .xlsx or .xlsm: "
            f"{template_path}"
        )

    output_dir = Path(output_directory).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = _resolve_output_filename(
        template_path,
        output_filename,
        project_name,
    )
    output_path = output_dir / filename

    if output_path.exists() and not overwrite:
        raise SensorMapTemplateError(
            f"Output file already exists: {output_path}"
        )

    sensor_sections = config["sensor_sections"]
    sensors_to_keep = parse_peripheral_sensor_scope(
        peripheral_sensor_scope,
        sensor_sections,
    )

    delete_rule = (
        config.get("matching", {})
        .get("delete_rule", "delete_if_not_present")
    )
    if delete_rule != "delete_if_not_present":
        raise SensorMapConfigError(
            'Only "delete_if_not_present" is currently supported as '
            '"matching.delete_rule".'
        )

    temporary_path = output_path.with_name(
        f".{output_path.stem}.tmp{output_path.suffix}"
    )

    try:
        shutil.copy2(template_path, temporary_path)

        keep_vba = template_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(
            temporary_path,
            keep_vba=keep_vba,
            keep_links=True,
        )

        try:
            worksheet = _select_worksheet(
                workbook,
                config.get("worksheet_name"),
            )
            detected_sections = find_sensor_sections(
                worksheet,
                config,
            )
            removed_sections = delete_unused_sensor_sections(
                worksheet,
                detected_sections,
                sensors_to_keep,
            )

            workbook.save(temporary_path)
        finally:
            workbook.close()

        temporary_path.replace(output_path)

    except PermissionError as exc:
        raise SensorMapTemplateError(
            "Permission denied while reading or writing the Sensor Map file. "
            "Confirm that the template/output file is not open in Excel and "
            "that the Public Link is writable."
        ) from exc
    except SensorMapError:
        raise
    except Exception as exc:
        LOGGER.exception("Unexpected Sensor Map generation failure.")
        raise SensorMapError(
            f"Unexpected Sensor Map generation failure: {exc}"
        ) from exc
    finally:
        if temporary_path.exists():
            try:
                temporary_path.unlink()
            except OSError:
                LOGGER.warning(
                    "Unable to remove temporary Sensor Map file: %s",
                    temporary_path,
                )

    configured_codes = {
        str(code).upper().strip() for code in sensor_sections
    }

    return {
        "success": True,
        "message": "Sensor Map generated successfully.",
        "template_path": str(template_path),
        "output_path": str(output_path),
        "worksheet": config.get("worksheet_name") or "active",
        "scope": peripheral_sensor_scope or "",
        "kept_sections": sorted(sensors_to_keep),
        "removed_sections": removed_sections,
        "configured_sections": sorted(configured_codes),
    }


if __name__ == "__main__":
    """
    Minimal local smoke test.

    Edit the three values below, then run:
        python PUMA_Server/services/sensormap_service.py
    """
    TEST_SCOPE = "UFS+PAS+PPS"
    TEST_OUTPUT_DIRECTORY = Path.cwd() / "sensormap_test_output"
    TEST_PROJECT_NAME = "Demo_Project"

    try:
        result = generate_sensor_map(
            peripheral_sensor_scope=TEST_SCOPE,
            output_directory=TEST_OUTPUT_DIRECTORY,
            project_name=TEST_PROJECT_NAME,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except SensorMapError as error:
        print(f"Sensor Map test failed: {error}")
